"""
report_exporter.py
Build a multi-sheet Excel workbook with portfolio summary, weights,
risk metrics, and price data.
"""

import io
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter


# ─── Colour constants ────────────────────────────────────────────────────────
DARK_FILL   = PatternFill("solid", fgColor="1A1A2E")
HEADER_FILL = PatternFill("solid", fgColor="00C9A7")
ALT_FILL    = PatternFill("solid", fgColor="16213E")
HEADER_FONT = Font(bold=True, color="000000", name="Calibri", size=11)
BODY_FONT   = Font(color="E0E0E0", name="Calibri", size=10)
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")
THIN        = Side(style="thin", color="2A2A4A")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header_row(ws, row_idx: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


def build_excel_report(
    tickers: list,
    prices: pd.DataFrame,
    returns: pd.DataFrame,
    ms_weights: np.ndarray,
    mv_weights: np.ndarray,
    ms_metrics: dict,
    mv_metrics: dict,
    sim_df: pd.DataFrame,
) -> bytes:
    """
    Create an Excel workbook and return as bytes for Streamlit download.
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ── Sheet 1: Summary ──────────────────────────────────────────────
        summary_rows = []
        for key in ms_metrics:
            summary_rows.append(
                {
                    "Metric": key,
                    "Max Sharpe": round(ms_metrics[key], 4),
                    "Min Volatility": round(mv_metrics[key], 4),
                }
            )
        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # ── Sheet 2: Portfolio Weights ────────────────────────────────────
        weights_df = pd.DataFrame(
            {
                "Ticker": tickers,
                "Max Sharpe Weight (%)": (ms_weights * 100).round(2),
                "Min Volatility Weight (%)": (mv_weights * 100).round(2),
            }
        )
        weights_df.to_excel(writer, sheet_name="Weights", index=False)

        # ── Sheet 3: Price Data ───────────────────────────────────────────
        prices.reset_index().to_excel(writer, sheet_name="Prices", index=False)

        # ── Sheet 4: Daily Returns ────────────────────────────────────────
        returns.reset_index().to_excel(writer, sheet_name="Returns", index=False)

        # ── Sheet 5: Correlation Matrix ───────────────────────────────────
        corr = returns.corr().round(4)
        corr.to_excel(writer, sheet_name="Correlation")

        # ── Sheet 6: Top Simulated Portfolios ────────────────────────────
        top50 = (
            sim_df.nlargest(50, "Sharpe")
            .round(4)
            .reset_index(drop=True)
        )
        top50.to_excel(writer, sheet_name="Top Portfolios", index=False)

    # ── Apply styling ────────────────────────────────────────────────────────
    output.seek(0)
    wb = load_workbook(output)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = "00C9A7"

        # Style header row
        _style_header_row(ws, 1, ws.max_column)

        # Style body
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = ALT_FILL if row_idx % 2 == 0 else DARK_FILL
            for cell in row:
                cell.fill = fill
                cell.font = BODY_FONT
                cell.alignment = LEFT
                cell.border = BORDER

        _auto_width(ws)

    final = io.BytesIO()
    wb.save(final)
    final.seek(0)
    return final
